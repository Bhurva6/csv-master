import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import requests
import re
import time
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import openpyxl.utils

# Configure the page
st.set_page_config(
    page_title="CSV Data Explorer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Main title
st.title("📊 CSV Operations")
st.markdown("---")

# Features section - Always visible
st.header("✨ Features")

col1, col2, col3, col4, col5 = st.columns(5)

with col1:
    if st.button("👨‍💼 Create Personas", use_container_width=True):
        st.session_state['feature'] = 'personas'
    st.markdown("""
    Generate detailed customer personas from your data
    """)

with col2:
    if st.button("📧 Verify Emails", use_container_width=True):
        st.session_state['feature'] = 'emails'
    st.markdown("""
    Advanced email validation with legitimacy scoring & Excel export
    """)

with col3:
    if st.button("📞 Verify Phone Numbers", use_container_width=True):
        st.session_state['feature'] = 'phones'
    st.markdown("""
    Advanced phone validation with format checking, region detection & Excel export
    """)

with col4:
    if st.button("🔍 Auto-Fill Missing Data", use_container_width=True):
        st.session_state['feature'] = 'autofill'
    st.markdown("""
    Automatically find and fill missing emails & phone numbers
    """)

with col5:
    if st.button("📈 Increase Reach", use_container_width=True):
        st.session_state['feature'] = 'increase_reach'
    st.markdown("""
    Find new leads by location and profession using web scraping
    """)

st.markdown("---")

# Sidebar for file upload
st.sidebar.header("Upload Your Data")
uploaded_file = st.sidebar.file_uploader(
    "Choose a CSV file",
    type="csv",
    help="Upload a CSV file to explore and visualize your data"
)

# Check for uploaded file or sample data
if uploaded_file is not None or 'uploaded_file' in st.session_state:
    # Handle both real uploads and sample data
    if uploaded_file is not None:
        file_data = uploaded_file
        filename = uploaded_file.name
    else:
        # Use sample data from session state
        file_data = st.session_state['uploaded_file']
        filename = st.session_state.get('uploaded_filename', 'sample_data.csv')
    
    try:
        # Read the CSV file
        if isinstance(file_data, bytes):
            import io
            df = pd.read_csv(io.BytesIO(file_data))
        else:
            df = pd.read_csv(file_data)
        
        # Store the DataFrame in session state for use in features
        st.session_state['df'] = df
        
        # Data preview
        st.subheader("🔍 Data Preview")
        st.dataframe(df.head(100), use_container_width=True)
        
        # Column information
        st.subheader("📊 Column Information")
        col_info = pd.DataFrame({
            'Column': df.columns,
            'Data Type': df.dtypes,
            'Non-Null Count': df.count(),
            'Null Count': df.isnull().sum(),
            'Unique Values': df.nunique()
        })
        st.dataframe(col_info, use_container_width=True)
        
        # Data visualization section
        st.header("📈 Data Visualization")
        
        # Select columns for visualization
        numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
        categorical_columns = df.select_dtypes(include=['object']).columns.tolist()
        
        if numeric_columns:
            st.subheader("Numeric Data Analysis")
            
            # Statistical summary
            st.write("**Statistical Summary:**")
            st.dataframe(df[numeric_columns].describe(), use_container_width=True)
            
            # Correlation heatmap
            if len(numeric_columns) > 1:
                st.write("**Correlation Heatmap:**")
                correlation_matrix = df[numeric_columns].corr()
                fig_heatmap = px.imshow(
                    correlation_matrix,
                    text_auto=True,
                    aspect="auto",
                    title="Correlation Matrix"
                )
                st.plotly_chart(fig_heatmap, use_container_width=True)
            
            # Distribution plots
            st.write("**Distribution Plots:**")
            selected_numeric = st.selectbox("Select a numeric column:", numeric_columns)
            if selected_numeric:
                fig_hist = px.histogram(
                    df, 
                    x=selected_numeric, 
                    title=f"Distribution of {selected_numeric}",
                    marginal="box"
                )
                st.plotly_chart(fig_hist, use_container_width=True)
        
        if categorical_columns:
            st.subheader("Categorical Data Analysis")
            selected_categorical = st.selectbox("Select a categorical column:", categorical_columns)
            if selected_categorical:
                # Value counts
                value_counts = df[selected_categorical].value_counts().head(20)
                
                col1, col2 = st.columns(2)
                with col1:
                    # Bar chart
                    fig_bar = px.bar(
                        x=value_counts.index,
                        y=value_counts.values,
                        title=f"Top Values in {selected_categorical}",
                        labels={'x': selected_categorical, 'y': 'Count'}
                    )
                    st.plotly_chart(fig_bar, use_container_width=True)
                
                with col2:
                    # Pie chart
                    fig_pie = px.pie(
                        values=value_counts.values,
                        names=value_counts.index,
                        title=f"Distribution of {selected_categorical}"
                    )
                    st.plotly_chart(fig_pie, use_container_width=True)
        
        # Custom visualization
        if len(numeric_columns) >= 2:
            st.subheader("Custom Scatter Plot")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                x_axis = st.selectbox("X-axis:", numeric_columns, key="x_axis")
            with col2:
                y_axis = st.selectbox("Y-axis:", numeric_columns, key="y_axis", index=1 if len(numeric_columns) > 1 else 0)
            with col3:
                color_by = st.selectbox("Color by:", ["None"] + categorical_columns, key="color_by")
            
            if x_axis and y_axis:
                fig_scatter = px.scatter(
                    df,
                    x=x_axis,
                    y=y_axis,
                    color=color_by if color_by != "None" else None,
                    title=f"{y_axis} vs {x_axis}",
                    hover_data=df.columns[:5].tolist()  # Show first 5 columns on hover
                )
                st.plotly_chart(fig_scatter, use_container_width=True)
        
        # Data filtering
        st.header("🔍 Data Filtering")
        st.write("Filter your data based on column values:")
        
        filters = {}
        for col in df.columns:
            if df[col].dtype in ['object']:
                unique_values = df[col].dropna().unique()
                if len(unique_values) <= 20:  # Only show filter for columns with <= 20 unique values
                    selected_values = st.multiselect(
                        f"Filter by {col}:",
                        options=unique_values,
                        key=f"filter_{col}"
                    )
                    if selected_values:
                        filters[col] = selected_values
            elif df[col].dtype in ['int64', 'float64']:
                min_val, max_val = float(df[col].min()), float(df[col].max())
                if min_val != max_val:
                    range_values = st.slider(
                        f"Filter {col} range:",
                        min_value=min_val,
                        max_value=max_val,
                        value=(min_val, max_val),
                        key=f"range_{col}"
                    )
                    if range_values != (min_val, max_val):
                        filters[col] = range_values
        
        # Apply filters
        filtered_df = df.copy()
        for col, values in filters.items():
            if df[col].dtype in ['object']:
                filtered_df = filtered_df[filtered_df[col].isin(values)]
            else:
                filtered_df = filtered_df[
                    (filtered_df[col] >= values[0]) & 
                    (filtered_df[col] <= values[1])
                ]
        
        if len(filters) > 0:
            st.write(f"**Filtered Data ({len(filtered_df)} rows)**")
            st.dataframe(filtered_df, use_container_width=True)
            
            # Download filtered data
            csv = filtered_df.to_csv(index=False)
            st.download_button(
                label="Download filtered data as CSV",
                data=csv,
                file_name="filtered_data.csv",
                mime="text/csv"
            )
    
    except Exception as e:
        st.error(f"Error reading the CSV file: {str(e)}")
        st.info("Please make sure your file is a valid CSV format.")

else:
    # Landing page when no file is uploaded
    st.info("👈 Please upload a CSV file using the sidebar to get started!")
    
    # Sample data section
    st.header("🎯 Try with Sample Data")
    
    if st.button("Generate Sample Data"):
        # Create sample data
        np.random.seed(42)
        sample_data = pd.DataFrame({
            'Name': [f'Person_{i}' for i in range(100)],
            'Age': np.random.randint(18, 65, 100),
            'Salary': np.random.normal(50000, 15000, 100),
            'Department': np.random.choice(['Engineering', 'Sales', 'Marketing', 'HR'], 100),
            'Experience': np.random.randint(0, 20, 100),
            'Performance_Score': np.random.uniform(1, 5, 100),
            'Gender': np.random.choice(['Male', 'Female', 'Non-binary', 'Prefer not to say'], 100),
            'City': np.random.choice(['New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix', 'Philadelphia', 'San Antonio', 'San Diego', 'Dallas', 'San Jose'], 100),
            'Email': [
                f"{f'Person_{i}'.lower().replace(' ', '.')}.{i}@example.com" if i % 10 != 0 else 
                f"user{i}@gmail.com" if i % 10 == 1 else
                f"contact{i}@yahoo.com" if i % 10 == 2 else
                f"info{i}@company.com" if i % 10 == 3 else
                f"test{i}@temp-mail.org" if i % 10 == 4 else
                f"person{i}@outlook.com" if i % 10 == 5 else
                f"employee{i}@business.net" if i % 10 == 6 else
                f"client{i}@mailinator.com" if i % 10 == 7 else
                "" if i % 10 == 8 else  # Empty email for testing
                f"invalid-email{i}"  # Invalid format for testing
                for i in range(100)
            ],
            'Phone': [
                f"+1-{np.random.randint(200,999)}-{np.random.randint(100,999)}-{np.random.randint(1000,9999)}" if i % 15 != 0 else
                f"({np.random.randint(200,999)}) {np.random.randint(100,999)}-{np.random.randint(1000,9999)}" if i % 15 == 1 else
                f"+44 {np.random.randint(20,99)} {np.random.randint(1000,9999)} {np.random.randint(1000,9999)}" if i % 15 == 2 else
                f"+91 {np.random.randint(7000000000,9999999999)}" if i % 15 == 3 else
                f"{np.random.randint(200,999)}-{np.random.randint(100,999)}-{np.random.randint(1000,9999)}" if i % 15 == 4 else
                f"+1-{np.random.randint(200,999)}-{np.random.randint(100,999)}-{np.random.randint(1000,9999)}, +44 {np.random.randint(20,99)} {np.random.randint(1000,9999)} {np.random.randint(1000,9999)}" if i % 15 == 5 else  # Multiple phones
                f"({np.random.randint(200,999)}) {np.random.randint(100,999)}-{np.random.randint(1000,9999)}; +91 {np.random.randint(7000000000,9999999999)}" if i % 15 == 6 else  # Multiple phones
                f"invalid-phone-{i}" if i % 15 == 7 else  # Invalid format
                f"123" if i % 15 == 8 else  # Too short
                "" if i % 15 == 9 else  # Empty
                f"+{np.random.randint(100,999)} {np.random.randint(100000,999999)}"  # International format
                for i in range(100)
            ]
        })
        
        st.success("Sample data generated!")
        st.dataframe(sample_data.head(10), use_container_width=True)
        
        # Store sample data in session state for immediate use
        st.session_state['sample_data'] = sample_data
        
        # Download sample data
        csv_sample = sample_data.to_csv(index=False)
        st.download_button(
            label="Download sample data",
            data=csv_sample,
            file_name="sample_data.csv",
            mime="text/csv"
        )
        
        # Button to use sample data immediately
        if st.button("Use This Sample Data Now"):
            # Convert to CSV bytes and simulate upload
            csv_bytes = sample_data.to_csv(index=False).encode()
            st.session_state['uploaded_file'] = csv_bytes
            st.session_state['uploaded_filename'] = 'sample_data.csv'
            st.rerun()

# Feature implementation - Always available when feature is selected
if 'feature' in st.session_state:
    st.markdown("---")

    if st.session_state['feature'] == 'personas':
        st.header("👨‍💼 Create Personas")

        # Persona creation logic
        if 'df' in st.session_state:
            try:
                # Use the DataFrame from session state
                df = st.session_state['df']

                # Check for relevant columns - expanded detection
                name_cols = [col for col in df.columns if any(keyword in col.lower() for keyword in ['name', 'full_name', 'first_name', 'last_name', 'customer_name', 'person_name', 'user_name'])]
                age_cols = [col for col in df.columns if any(keyword in col.lower() for keyword in ['age', 'age_years', 'person_age', 'user_age', 'years_old'])]
                gender_cols = [col for col in df.columns if any(keyword in col.lower() for keyword in ['gender', 'sex', 'male', 'female', 'person_gender', 'user_gender'])]
                location_cols = [col for col in df.columns if any(keyword in col.lower() for keyword in ['city', 'state', 'country', 'location', 'address', 'region', 'province', 'town', 'place'])]

                if name_cols or age_cols or gender_cols or location_cols:
                    found_cols = []
                    if name_cols:
                        found_cols.append(f"Name: {', '.join(name_cols)}")
                    if age_cols:
                        found_cols.append(f"Age: {', '.join(age_cols)}")
                    if gender_cols:
                        found_cols.append(f"Gender: {', '.join(gender_cols)}")
                    if location_cols:
                        found_cols.append(f"Location: {', '.join(location_cols)}")
                    
                    st.success(f"Found relevant columns for persona creation: {', '.join(found_cols)}")

                    # Allow user to select number of personas
                    max_personas = min(10, len(df))  # Limit to 10 or available rows
                    num_personas = st.slider("Number of personas to generate:", min_value=1, max_value=max_personas, value=min(5, max_personas))

                    # Generate personas from random sample
                    if len(df) > num_personas:
                        # Randomly sample rows for diversity
                        sample_indices = np.random.choice(len(df), size=num_personas, replace=False)
                    else:
                        sample_indices = range(len(df))

                    # Generate personas
                    personas = []
                    for i, idx in enumerate(sample_indices):
                        persona = {
                            'Name': f"Persona {i+1}",
                            'Age': "Unknown",
                            'Gender': "Unknown", 
                            'Location': "Unknown"
                        }
                        
                        if name_cols:
                            persona['Name'] = df[name_cols[0]].iloc[idx] if pd.notna(df[name_cols[0]].iloc[idx]) else f"Persona {i+1}"
                        
                        if age_cols:
                            persona['Age'] = df[age_cols[0]].iloc[idx] if pd.notna(df[age_cols[0]].iloc[idx]) else "Unknown"
                        
                        if gender_cols:
                            persona['Gender'] = df[gender_cols[0]].iloc[idx] if pd.notna(df[gender_cols[0]].iloc[idx]) else "Unknown"
                        
                        if location_cols:
                            persona['Location'] = df[location_cols[0]].iloc[idx] if pd.notna(df[location_cols[0]].iloc[idx]) else "Unknown"

                        personas.append(persona)

                    # Display personas
                    for i, persona in enumerate(personas):
                        with st.expander(f"**{persona['Name']}**"):
                            col1, col2 = st.columns(2)
                            with col1:
                                st.write(f"**Age:** {persona['Age']}")
                                st.write(f"**Gender:** {persona['Gender']}")
                            with col2:
                                st.write(f"**Location:** {persona['Location']}")
                                st.write("**Characteristics:** Based on data analysis")

                    # Download personas
                    persona_df = pd.DataFrame(personas)
                    csv_personas = persona_df.to_csv(index=False)
                    st.download_button(
                        label="Download Personas as CSV",
                        data=csv_personas,
                        file_name="personas.csv",
                        mime="text/csv"
                    )
                else:
                    st.warning("No relevant columns found for persona creation. Please ensure your CSV has columns containing keywords like 'name', 'age', 'gender', 'sex', 'city', 'state', 'country', or 'location'.")
                    st.info(f"Available columns in your CSV: {', '.join(df.columns.tolist())}")
                    st.info("If your columns have different names, you may need to rename them or contact support.")

            except Exception as e:
                st.error(f"Error processing data for personas: {str(e)}")
        else:
            st.info("Please upload a CSV file first to create personas.")

    elif st.session_state['feature'] == 'emails':
        st.header("📧 Verify Emails")

        # Email verification logic - works with uploaded CSV or sample data
        if 'df' in st.session_state:
            try:
                # Use the DataFrame from session state
                df = st.session_state['df']

                # Find email columns
                email_cols = [col for col in df.columns if any(keyword in col.lower() for keyword in ['email', 'mail', 'e-mail', 'email_address'])]

                if email_cols:
                    st.success(f"Found email column(s): {', '.join(email_cols)}")

                    # Select which email column to verify
                    selected_email_col = st.selectbox("Select email column to verify:", email_cols)

                    if st.button("Verify Emails in CSV"):
                        with st.spinner("Verifying emails..."):
                            import re

                            # Import dns.resolver for MX record checking
                            import dns.resolver

                            # Enhanced email validation
                            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'

                            # Common disposable/temporary email domains
                            disposable_domains = [
                                '10minutemail.com', 'guerrillamail.com', 'mailinator.com',
                                'temp-mail.org', 'throwaway.email', 'yopmail.com',
                                'maildrop.cc', 'tempail.com', 'dispostable.com'
                            ]

                            # Function to check if domain has valid MX records
                            def has_valid_mx_record(domain):
                                try:
                                    mx_records = dns.resolver.resolve(domain, 'MX')
                                    return len(mx_records) > 0
                                except:
                                    return False

                            results = []
                            for idx, email_data in enumerate(df[selected_email_col]):
                                if pd.isna(email_data) or str(email_data).strip() == '':
                                    results.append({
                                        'Row': idx + 1,
                                        'Original Data': email_data if pd.notna(email_data) else '',
                                        'Email': '',
                                        'Status': 'Empty',
                                        'Notes': 'Missing email address'
                                    })
                                    continue

                                # Handle multiple emails in a single cell
                                email_string = str(email_data).strip()
                                
                                # Split by common delimiters: comma, semicolon, newline, pipe, slash
                                emails_in_cell = re.split(r'[;,|\n/]+', email_string)
                                emails_in_cell = [e.strip() for e in emails_in_cell if e.strip()]

                                if not emails_in_cell:
                                    results.append({
                                        'Row': idx + 1,
                                        'Original Data': email_string,
                                        'Email': '',
                                        'Status': 'Empty',
                                        'Notes': 'No emails found after parsing'
                                    })
                                    continue

                                # Process each email in the cell
                                for email_idx, email in enumerate(emails_in_cell):
                                    email = email.lower()
                                    is_valid_format = bool(re.match(email_pattern, email))

                                if is_valid_format:
                                    domain = email.split('@')[1]
                                    is_disposable = domain in disposable_domains
                                    is_deliverable = not any(x in email for x in ['test', 'example', 'invalid', 'fake'])
                                    has_mx = has_valid_mx_record(domain)

                                    if is_deliverable and not is_disposable and has_mx:
                                        status = 'Valid (Deliverable)'
                                        notes = f'Valid format, {domain} domain with MX records'
                                    elif is_disposable:
                                        status = 'Invalid (Disposable)'
                                        notes = f'Disposable email domain: {domain}'
                                    elif not has_mx:
                                        status = 'Invalid (Unreachable)'
                                        notes = f'Domain {domain} has no valid MX records'
                                    else:
                                        status = 'Invalid (Undeliverable)'
                                        notes = f'Potentially undeliverable: {domain}'
                                else:
                                    status = 'Invalid (Format)'
                                    notes = 'Invalid email format'

                                results.append({
                                    'Row': idx + 1,
                                    'Original Data': email_string if email_idx == 0 else '',  # Show original data only for first email
                                    'Email': email,
                                    'Status': status,
                                    'Notes': f'{notes} (Email {email_idx + 1} of {len(emails_in_cell)})' if len(emails_in_cell) > 1 else notes
                                })

                            results_df = pd.DataFrame(results)

                            # Summary metrics
                            st.subheader("📊 Email Verification Summary")
                            
                            total_emails = len(results)
                            valid_deliverable = sum(1 for r in results if r['Status'] == 'Valid (Deliverable)')
                            invalid_emails = sum(1 for r in results if r['Status'].startswith('Invalid'))
                            empty_emails = sum(1 for r in results if r['Status'] == 'Empty')

                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("Total Emails", total_emails)
                            with col2:
                                st.metric("Valid (Deliverable)", valid_deliverable)
                            with col3:
                                st.metric("Invalid Emails", invalid_emails)
                            with col4:
                                st.metric("Empty Emails", empty_emails)

                            # Success rate
                            if total_emails > 0:
                                success_rate = (valid_deliverable / total_emails) * 100
                                st.success(f"✅ **{success_rate:.1f}%** of emails are valid and deliverable!")

                            # Visual representation
                            status_counts = results_df['Status'].value_counts()
                            fig_pie = px.pie(
                                values=status_counts.values,
                                names=status_counts.index,
                                title="Email Verification Status Distribution",
                                color_discrete_sequence=['#00ff00', '#ff0000', '#ffff00', '#ff9800']
                            )
                            st.plotly_chart(fig_pie, use_container_width=True)

                            # Show results table
                            st.subheader("📋 Detailed Results")
                            st.dataframe(results_df, use_container_width=True)

                            # Create Excel file for download
                            from io import BytesIO
                            import openpyxl
                            from openpyxl.styles import PatternFill, Font

                            wb = openpyxl.Workbook()
                            ws = wb.active
                            ws.title = "Email Verification Results"

                            # Add summary at the top
                            ws['A1'] = "Email Verification Summary"
                            ws['A1'].font = Font(bold=True, size=14)
                            
                            ws['A3'] = "Total Emails:"
                            ws['B3'] = total_emails
                            ws['A4'] = "Valid (Deliverable):"
                            ws['B4'] = valid_deliverable
                            ws['A5'] = "Invalid Emails:"
                            ws['B5'] = invalid_emails
                            ws['A6'] = "Empty Emails:"
                            ws['B6'] = empty_emails
                            ws['A7'] = f"Success Rate: {success_rate:.1f}%" if total_emails > 0 else "Success Rate: N/A"

                            # Add detailed results starting from row 10
                            start_row = 10
                            headers = list(results_df.columns)
                            for col_num, header in enumerate(headers, 1):
                                cell = ws.cell(row=start_row, column=col_num, value=header)
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                            # Add data with conditional formatting
                            for row_num, row_data in enumerate(results_df.values, start_row + 1):
                                for col_num, value in enumerate(row_data, 1):
                                    cell = ws.cell(row=row_num, column=col_num, value=str(value))

                                    # Color coding based on status
                                    if col_num == 3:  # Status column
                                        if 'Valid' in str(value):
                                            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                                        elif 'Invalid' in str(value):
                                            cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                                        elif 'Empty' in str(value):
                                            cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

                            # Auto-adjust column widths
                            for column in ws.columns:
                                max_length = 0
                                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                ws.column_dimensions[column_letter].width = adjusted_width

                            # Save to BytesIO
                            excel_buffer = BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)

                            # Download button
                            st.download_button(
                                label="📥 Download Verification Results (Excel)",
                                data=excel_buffer,
                                file_name="email_verification_results.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                else:
                    st.warning("No email columns found in your CSV. Please ensure your file has columns containing 'email', 'mail', 'e-mail', or 'email_address' in the name.")
                    st.info(f"Available columns in your CSV: {', '.join(df.columns.tolist())}")
                    st.info("If your email column has a different name, you may need to rename it or contact support.")

            except Exception as e:
                st.error(f"Error processing CSV for email verification: {str(e)}")
        else:
            st.info("Please upload a CSV file first to verify emails.")

    elif st.session_state['feature'] == 'phones':
        st.header("📞 Verify Phone Numbers")

        # Phone verification logic - works with uploaded CSV or sample data
        if 'df' in st.session_state:
            try:
                # Use the DataFrame from session state
                df = st.session_state['df']

                # Find phone columns
                phone_cols = [col for col in df.columns if any(keyword in col.lower() for keyword in ['phone', 'mobile', 'tel', 'contact', 'cell', 'telephone', 'number'])]

                if phone_cols:
                    st.success(f"Found phone column(s): {', '.join(phone_cols)}")

                    # Select which phone column to verify
                    selected_phone_col = st.selectbox("Select phone column to verify:", phone_cols)

                    if st.button("Verify Phone Numbers in CSV"):
                        with st.spinner("Verifying phone numbers..."):
                            import re

                            results = []
                            for idx, phone_data in enumerate(df[selected_phone_col]):
                                if pd.isna(phone_data) or str(phone_data).strip() == '':
                                    results.append({
                                        'Row': idx + 1,
                                        'Original Data': phone_data if pd.notna(phone_data) else '',
                                        'Phone Number': '',
                                        'Status': 'Empty',
                                        'Notes': 'Missing phone number'
                                    })
                                    continue

                                # Handle multiple phone numbers in a single cell
                                phone_string = str(phone_data).strip()
                                
                                # Split by common delimiters: comma, semicolon, newline, pipe, slash
                                phones_in_cell = re.split(r'[;,|\n/]+', phone_string)
                                phones_in_cell = [p.strip() for p in phones_in_cell if p.strip()]

                                if not phones_in_cell:
                                    results.append({
                                        'Row': idx + 1,
                                        'Original Data': phone_string,
                                        'Phone Number': '',
                                        'Status': 'Empty',
                                        'Notes': 'No phone numbers found after parsing'
                                    })
                                    continue

                                # Process each phone number in the cell
                                for phone_idx, phone in enumerate(phones_in_cell):
                                    # Remove all non-digit characters except + for international codes
                                    clean_phone = re.sub(r'[^\d+]', '', phone)

                                    # Basic validation patterns
                                    us_pattern = r'^\+?1?\d{10}$'
                                    international_pattern = r'^\+\d{1,4}\d{6,14}$'

                                    is_us_format = bool(re.match(us_pattern, clean_phone))
                                    is_international = bool(re.match(international_pattern, clean_phone))

                                    # Determine country/region and phone type
                                    if clean_phone.startswith('+1') or (len(clean_phone) == 10 and not clean_phone.startswith('+')):
                                        region = "US/Canada"
                                        # US/Canada: Landline prefixes typically don't start with 1
                                        is_landline = not (clean_phone[-10] in '2345')
                                    elif clean_phone.startswith('+44'):
                                        region = "UK"
                                        # UK: Landline numbers typically start with 01 or 02
                                        is_landline = clean_phone[3:5] in ['01', '02']
                                    elif clean_phone.startswith('+91'):
                                        region = "India"
                                        # India: Landline numbers are typically 8 digits and start with area codes
                                        is_landline = len(clean_phone[3:]) <= 8
                                    elif clean_phone.startswith('+61'):
                                        region = "Australia"
                                        # Australian landlines start with 02-08
                                        is_landline = clean_phone[3:4] in ['2', '3', '7', '8']
                                    elif clean_phone.startswith('+86'):
                                        region = "China"
                                        is_landline = len(clean_phone[3:]) < 11
                                    elif clean_phone.startswith('+81'):
                                        region = "Japan"
                                        is_landline = not clean_phone[3] in ['7', '8', '9']
                                    elif clean_phone.startswith('+49'):
                                        region = "Germany"
                                        is_landline = not clean_phone[3] in ['1', '5']
                                    elif clean_phone.startswith('+33'):
                                        region = "France"
                                        is_landline = clean_phone[3] == '1'
                                    elif clean_phone.startswith('+'):
                                        region = "International"
                                        is_landline = False  # Default to mobile for unknown international
                                    else:
                                        region = "Unknown"
                                        is_landline = False

                                    # Check if number is likely to be WhatsApp enabled
                                    # Most mobile numbers can use WhatsApp, so we'll assume mobile numbers are WhatsApp-enabled
                                    is_whatsapp = is_valid and not is_landline

                                    is_valid = is_us_format or is_international

                                    results.append({
                                        'Row': idx + 1,
                                        'Original Data': phone_string if phone_idx == 0 else '',
                                        'Phone Number': phone,
                                        'Cleaned': clean_phone,
                                        'Format Valid': '✅ Valid' if is_valid else '❌ Invalid',
                                        'Region': region,
                                        'Type': 'Landline' if is_landline else 'Mobile',
                                        'WhatsApp': '✅ Yes' if is_whatsapp else '❌ No',
                                        'Status': 'Valid' if is_valid else 'Invalid',
                                        'Notes': f'Phone {phone_idx + 1} of {len(phones_in_cell)} in cell' if len(phones_in_cell) > 1 else ''
                                    })

                            results_df = pd.DataFrame(results)

                            # Summary metrics
                            st.subheader("📊 Phone Verification Summary")
                            
                            total_phones = len(results)
                            valid_phones = sum(1 for r in results if r['Status'] == 'Valid')
                            invalid_phones = sum(1 for r in results if r['Status'] == 'Invalid')
                            empty_entries = sum(1 for r in results if r['Status'] == 'Empty')

                            # Add WhatsApp and phone type counts
                            whatsapp_numbers = sum(1 for r in results if r['WhatsApp'] == '✅ Yes')
                            landline_numbers = sum(1 for r in results if r['Status'] == 'Valid' and r['Type'] == 'Landline')
                            mobile_numbers = sum(1 for r in results if r['Status'] == 'Valid' and r['Type'] == 'Mobile')

                            col1, col2 = st.columns(2)
                            with col1:
                                st.metric("Total Phone Numbers", total_phones)
                                st.metric("Valid Numbers", valid_phones)
                                st.metric("Invalid Numbers", invalid_phones)
                            with col2:
                                st.metric("WhatsApp Enabled", whatsapp_numbers)
                                st.metric("Mobile Numbers", mobile_numbers)
                                st.metric("Landline Numbers", landline_numbers)

                            # Add separate sections for Mobile and Landline numbers
                            if mobile_numbers > 0:
                                st.subheader("📱 Mobile Numbers")
                                mobile_df = results_df[results_df['Type'] == 'Mobile']
                                st.dataframe(mobile_df, use_container_width=True)

                            if landline_numbers > 0:
                                st.subheader("☎️ Landline Numbers")
                                landline_df = results_df[results_df['Type'] == 'Landline']
                                st.dataframe(landline_df, use_container_width=True)

                            # Success rate
                            if total_phones > 0:
                                success_rate = (valid_phones / total_phones) * 100
                                st.success(f"✅ **{success_rate:.1f}%** of phone numbers are valid!")

                            # Visual representation
                            status_counts = results_df['Status'].value_counts()
                            fig_pie = px.pie(
                                values=status_counts.values,
                                names=status_counts.index,
                                title="Phone Verification Status Distribution",
                                color_discrete_sequence=['#00ff00', '#ff0000', '#ffff00', '#ff9800']
                            )
                            st.plotly_chart(fig_pie, use_container_width=True)

                            # Show results table
                            st.subheader("📋 Detailed Results")
                            st.dataframe(results_df, use_container_width=True)

                            # Create Excel file for download
                            from io import BytesIO
                            import openpyxl
                            from openpyxl.styles import PatternFill, Font

                            wb = openpyxl.Workbook()
                            ws = wb.active
                            ws.title = "Phone Verification Results"

                            # Add summary at the top
                            ws['A1'] = "Phone Verification Summary"
                            ws['A1'].font = Font(bold=True, size=14)
                            
                            ws['A3'] = "Total Phone Numbers:"
                            ws['B3'] = total_phones
                            ws['A4'] = "Valid Numbers:"
                            ws['B4'] = valid_phones
                            ws['A5'] = "Invalid Numbers:"
                            ws['B5'] = invalid_phones
                            ws['A6'] = "Empty Entries:"
                            ws['B6'] = empty_entries
                            ws['A7'] = f"Success Rate: {success_rate:.1f}%" if total_phones > 0 else "Success Rate: N/A"
                            ws['A8'] = "WhatsApp Enabled:"
                            ws['B8'] = whatsapp_numbers
                            ws['A9'] = "Mobile Numbers:"
                            ws['B9'] = mobile_numbers
                            ws['A10'] = "Landline Numbers:"
                            ws['B10'] = landline_numbers

                            # Add detailed results starting from row 10
                            start_row = 10
                            headers = list(results_df.columns)
                            for col_num, header in enumerate(headers, 1):
                                cell = ws.cell(row=start_row, column=col_num, value=header)
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                            # Add data with conditional formatting
                            for row_num, row_data in enumerate(results_df.values, start_row + 1):
                                for col_num, value in enumerate(row_data, 1):
                                    cell = ws.cell(row=row_num, column=col_num, value=str(value))

                                    # Color coding based on status
                                    if col_num == 6:  # Status column
                                        if 'Valid' in str(value):
                                            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                                        elif 'Invalid' in str(value):
                                            cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                                        elif 'Empty' in str(value):
                                            cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

                            # Auto-adjust column widths
                            for column in ws.columns:
                                max_length = 0
                                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                ws.column_dimensions[column_letter].width = adjusted_width

                            # Save to BytesIO
                            excel_buffer = BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)

                            # Download button
                            st.download_button(
                                label="📥 Download Verification Results (Excel)",
                                data=excel_buffer,
                                file_name="phone_verification_results.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                else:
                    st.warning("No phone columns found in your CSV. Please ensure your file has columns containing 'phone', 'mobile', 'tel', 'contact', 'cell', 'telephone', or 'number' in the name.")
                    st.info(f"Available columns in your CSV: {', '.join(df.columns.tolist())}")
                    st.info("If your phone column has a different name, you may need to rename it or contact support.")

            except Exception as e:
                st.error(f"Error processing CSV for phone verification: {str(e)}")
        else:
            st.info("Please upload a CSV file first to verify phone numbers.")
    
    elif st.session_state['feature'] == 'autofill':
        st.header("🔍 Auto-Fill Missing Data")
        st.info("This feature will search and fill missing or invalid emails and phone numbers based on available data.")

        if 'df' in st.session_state:
            try:
                df = st.session_state['df']

                # Find relevant columns
                name_cols = [col for col in df.columns if any(keyword in col.lower() for keyword in ['name', 'full_name', 'first_name', 'last_name', 'customer_name', 'person_name'])]
                email_cols = [col for col in df.columns if any(keyword in col.lower() for keyword in ['email', 'mail', 'e-mail'])]
                phone_cols = [col for col in df.columns if any(keyword in col.lower() for keyword in ['phone', 'mobile', 'tel', 'contact'])]
                company_cols = [col for col in df.columns if any(keyword in col.lower() for keyword in ['company', 'organization', 'business', 'employer'])]

                if name_cols or email_cols or phone_cols:
                    st.success(f"Found columns for auto-fill analysis")

                    # Configuration
                    st.subheader("⚙️ Configuration")
                    col1, col2 = st.columns(2)

                    with col1:
                        selected_name_col = st.selectbox("Name Column:", ["None"] + name_cols) if name_cols else "None"
                        selected_email_col = st.selectbox("Email Column:", ["None"] + email_cols) if email_cols else "None"

                    with col2:
                        selected_phone_col = st.selectbox("Phone Column:", ["None"] + phone_cols) if phone_cols else "None"
                        selected_company_col = st.selectbox("Company Column (optional):", ["None"] + company_cols) if company_cols else "None"

                    # Hardcoded API keys
                    google_api_key = "AIzaSyDummyKeyForDemoPurposesOnly"  # Replace with your actual Google API key
                    google_search_engine_id = "01234567890:abcdefghijk"  # Replace with your actual search engine ID
                    hunter_api_key = "demo_hunter_api_key_12345"  # Replace with your actual Hunter.io API key

                    search_method = "Real APIs (Production)"

                    if st.button("🚀 Start Auto-Fill Process"):
                        with st.spinner("Searching and filling missing data..."):

                            def simulate_search(name, company=None):
                                """Simulate web search results for demo purposes"""
                                time.sleep(0.05)  # Simulate search delay

                                if name and str(name) != "Unknown" and not pd.isna(name):
                                    # Generate plausible email based on name
                                    name_str = str(name).lower().strip()
                                    name_parts = name_str.split()

                                    if len(name_parts) >= 2:
                                        # Create realistic email formats
                                        formats = [
                                            f"{name_parts[0]}.{name_parts[-1]}@example.com",
                                            f"{name_parts[0][0]}{name_parts[-1]}@company.com",
                                            f"{name_parts[0]}_{name_parts[-1]}@business.net"
                                        ]
                                        email = np.random.choice(formats)
                                    else:
                                        email = f"{name_parts[0]}@example.com"

                                    # Generate phone
                                    phone = f"+1-{np.random.randint(200,999)}-{np.random.randint(100,999)}-{np.random.randint(1000,9999)}"

                                    # Simulate success rate (80% success)
                                    if np.random.random() < 0.8:
                                        return {
                                            'email': email,
                                            'phone': phone,
                                            'source': 'Simulated Web Search',
                                            'confidence': np.random.choice(['High', 'Medium', 'Low'], p=[0.5, 0.3, 0.2])
                                        }

                                return None

                            def search_with_google(name, company=None, api_key=None, engine_id=None):
                                """Search Google for contact information"""
                                if not api_key or not engine_id:
                                    return None

                                try:
                                    query = f"{name}"
                                    if company and company != "Unknown":
                                        query += f" {company}"
                                    query += " email phone contact"

                                    url = f"https://www.googleapis.com/customsearch/v1?key={api_key}&cx={engine_id}&q={requests.utils.quote(query)}&num=5"
                                    response = requests.get(url, timeout=10)

                                    if response.status_code == 200:
                                        data = response.json()
                                        email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
                                        phone_pattern = r'\+?1?\d{10,15}'

                                        emails = []
                                        phones = []

                                        for item in data.get('items', [])[:5]:
                                            snippet = item.get('snippet', '')
                                            title = item.get('title', '')
                                            content = snippet + ' ' + title

                                            emails.extend(re.findall(email_pattern, content))
                                            phones.extend(re.findall(phone_pattern, content))

                                        if emails or phones:
                                            return {
                                                'email': emails[0] if emails else None,
                                                'phone': phones[0] if phones else None,
                                                'source': 'Google Custom Search',
                                                'confidence': 'Medium'
                                            }
                                except Exception as e:
                                    st.warning(f"Google Search error: {str(e)}")

                                return None

                            def search_with_hunter(name, company=None, api_key=None):
                                """Search Hunter.io for email addresses"""
                                if not api_key:
                                    return None

                                try:
                                    # Try domain search first if company is available
                                    if company and company != "Unknown":
                                        domain = company.lower().replace(' ', '').replace(',', '').replace('.', '')
                                        url = f"https://api.hunter.io/v2/domain-search?domain={domain}&api_key={api_key}"
                                        response = requests.get(url, timeout=10)

                                        if response.status_code == 200:
                                            data = response.json()
                                            emails = data.get('data', {}).get('emails', [])

                                            if emails:
                                                # Find email that might match the name
                                                name_parts = str(name).lower().split()
                                                for email_data in emails:
                                                    email = email_data.get('value', '')
                                                    if name_parts:
                                                        first_name = name_parts[0]
                                                        if first_name in email.lower():
                                                            return {
                                                                'email': email,
                                                                'phone': None,
                                                                'source': 'Hunter.io',
                                                                'confidence': 'High'
                                                            }

                                    # Fallback to email finder
                                    if '@' not in str(name):
                                        url = f"https://api.hunter.io/v2/email-finder?full_name={requests.utils.quote(str(name))}&api_key={api_key}"
                                        if company and company != "Unknown":
                                            url += f"&company={requests.utils.quote(str(company))}"

                                        response = requests.get(url, timeout=10)

                                        if response.status_code == 200:
                                            data = response.json()
                                            email = data.get('data', {}).get('email')

                                            if email:
                                                return {
                                                    'email': email,
                                                    'phone': None,
                                                    'source': 'Hunter.io',
                                                    'confidence': 'High'
                                                }

                                except Exception as e:
                                    st.warning(f"Hunter.io error: {str(e)}")

                                return None

                            def perform_real_search(name, company=None):
                                """Perform search using available APIs"""
                                results = []

                                # Try Hunter.io first (best for emails)
                                if hunter_api_key:
                                    result = search_with_hunter(name, company, hunter_api_key)
                                    if result:
                                        return result

                                # Try Google as fallback
                                if google_api_key and google_search_engine_id:
                                    result = search_with_google(name, company, google_api_key, google_search_engine_id)
                                    if result:
                                        return result

                                return None

                            # Email validation function
                            def is_valid_email(email):
                                if pd.isna(email) or str(email).strip() == '':
                                    return False
                                email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                                return bool(re.match(email_pattern, str(email).strip()))

                            # Phone validation function
                            def is_valid_phone(phone):
                                if pd.isna(phone) or str(phone).strip() == '':
                                    return False
                                clean_phone = re.sub(r'[^\d+]', '', str(phone))
                                return len(clean_phone) >= 10

                            results = []
                            progress_bar = st.progress(0)
                            status_text = st.empty()

                            for idx in range(len(df)):
                                status_text.text(f"Processing row {idx + 1} of {len(df)}...")
                                progress_bar.progress((idx + 1) / len(df))

                                row_data = {
                                    'Row': idx + 1,
                                    'Name': df[selected_name_col].iloc[idx] if selected_name_col != "None" else "Unknown",
                                    'Original Email': df[selected_email_col].iloc[idx] if selected_email_col != "None" else "",
                                    'Original Phone': df[selected_phone_col].iloc[idx] if selected_phone_col != "None" else "",
                                    'Company': df[selected_company_col].iloc[idx] if selected_company_col != "None" else "",
                                    'New Email': "",
                                    'New Phone': "",
                                    'Email Status': "",
                                    'Phone Status': "",
                                    'Search Source': "",
                                    'Confidence': "",
                                    'Notes': ""
                                }

                                # Check if email/phone needs filling
                                needs_email = not is_valid_email(row_data['Original Email'])
                                needs_phone = not is_valid_phone(row_data['Original Phone'])

                                if needs_email or needs_phone:
                                    # Perform search
                                    search_result = perform_real_search(row_data['Name'], row_data['Company'])

                                    if search_result:
                                        if needs_email and search_result.get('email'):
                                            row_data['New Email'] = search_result['email']
                                            row_data['Email Status'] = '✅ Found'
                                            row_data['Notes'] += "Email found via web search. "
                                        elif needs_email:
                                            row_data['Email Status'] = '❌ Not Found'
                                            row_data['Notes'] += "Email not found in search results. "

                                        if needs_phone and search_result.get('phone'):
                                            row_data['New Phone'] = search_result['phone']
                                            row_data['Phone Status'] = '✅ Found'
                                            row_data['Notes'] += "Phone found via web search. "
                                        elif needs_phone:
                                            row_data['Phone Status'] = '❌ Not Found'
                                            row_data['Notes'] += "Phone not found in search results. "

                                        row_data['Search Source'] = search_result['source']
                                        row_data['Confidence'] = search_result['confidence']
                                    else:
                                        if needs_email:
                                            row_data['Email Status'] = '❌ Not Found'
                                            row_data['Notes'] += "No search results found. "
                                        if needs_phone:
                                            row_data['Phone Status'] = '❌ Not Found'
                                            row_data['Notes'] += "No search results found. "
                                else:
                                    row_data['Email Status'] = '✓ Already Valid'
                                    row_data['Phone Status'] = '✓ Already Valid'
                                    row_data['Notes'] = 'No action needed - data already valid'

                                results.append(row_data)

                            progress_bar.empty()
                            status_text.empty()

                            results_df = pd.DataFrame(results)

                            # Summary
                            st.subheader("📊 Auto-Fill Summary")

                            total_rows = len(results)
                            emails_found = sum(1 for r in results if r['Email Status'] == '✅ Found')
                            phones_found = sum(1 for r in results if r['Phone Status'] == '✅ Found')
                            emails_not_found = sum(1 for r in results if r['Email Status'] == '❌ Not Found')
                            phones_not_found = sum(1 for r in results if r['Phone Status'] == '❌ Not Found')
                            already_valid = sum(1 for r in results if 'already valid' in r['Notes'].lower())

                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Total Rows Processed", total_rows)
                                st.metric("Already Valid", already_valid)
                            with col2:
                                st.metric("Emails Found", emails_found)
                                st.metric("Emails Not Found", emails_not_found)
                            with col3:
                                st.metric("Phones Found", phones_found)
                                st.metric("Phones Not Found", phones_not_found)

                            # Show results
                            st.subheader("📋 Detailed Results")
                            st.dataframe(results_df, use_container_width=True)

                            # Create updated DataFrame
                            updated_df = df.copy()
                            if selected_email_col != "None":
                                for idx, row in results_df.iterrows():
                                    if row['New Email']:
                                        updated_df.at[idx, selected_email_col] = row['New Email']

                            if selected_phone_col != "None":
                                for idx, row in results_df.iterrows():
                                    if row['New Phone']:
                                        updated_df.at[idx, selected_phone_col] = row['New Phone']

                            # Create Excel with multiple sheets
                            wb = openpyxl.Workbook()

                            # Summary Sheet
                            ws_summary = wb.active
                            ws_summary.title = "Summary"
                            ws_summary['A1'] = "Auto-Fill Summary Report"
                            ws_summary['A1'].font = Font(bold=True, size=14)
                            ws_summary.merge_cells('A1:B1')

                            ws_summary['A3'] = "Total Rows:"
                            ws_summary['B3'] = total_rows
                            ws_summary['A4'] = "Emails Found:"
                            ws_summary['B4'] = emails_found
                            ws_summary['A5'] = "Phones Found:"
                            ws_summary['B5'] = phones_found
                            ws_summary['A6'] = "Emails Not Found:"
                            ws_summary['B6'] = emails_not_found
                            ws_summary['A7'] = "Phones Not Found:"
                            ws_summary['B7'] = phones_not_found
                            ws_summary['A8'] = "Already Valid:"
                            ws_summary['B8'] = already_valid

                            # Style summary
                            for row in range(3, 9):
                                ws_summary[f'A{row}'].font = Font(bold=True)

                            # Search Results Sheet
                            ws_results = wb.create_sheet("Search Results")
                            headers = list(results_df.columns)
                            for col_num, header in enumerate(headers, 1):
                                cell = ws_results.cell(row=1, column=col_num, value=header)
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.alignment = Alignment(horizontal="center")

                            for row_num, row_data in enumerate(results_df.values, 2):
                                for col_num, value in enumerate(row_data, 1):
                                    cell = ws_results.cell(row=row_num, column=col_num, value=str(value))

                                    # Color coding
                                    if '✅' in str(value):
                                        cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                                    elif '❌' in str(value):
                                        cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                                    elif '✓' in str(value):
                                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

                            # Updated Data Sheet
                            ws_updated = wb.create_sheet("Updated Data")
                            headers = list(updated_df.columns)
                            for col_num, header in enumerate(headers, 1):
                                cell = ws_updated.cell(row=1, column=col_num, value=header)
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.alignment = Alignment(horizontal="center")

                            for row_num, row_data in enumerate(updated_df.values, 2):
                                for col_num, value in enumerate(row_data, 1):
                                    ws_updated.cell(row=row_num, column=col_num, value=str(value))

                            # Auto-adjust column widths for all sheets
                            for ws in [ws_summary, ws_results, ws_updated]:
                                for column in ws.columns:
                                    max_length = 0
                                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                                    for cell in column:
                                        try:
                                            if cell.value and len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = min(max_length + 2, 50)
                                    ws.column_dimensions[column_letter].width = adjusted_width

                            # Save to BytesIO
                            excel_buffer = BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)

                            # Download buttons
                            col1, col2 = st.columns(2)
                            with col1:
                                st.download_button(
                                    label="📥 Download Complete Report (Excel with 3 sheets)",
                                    data=excel_buffer,
                                    file_name="autofill_results.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )

                            with col2:
                                # Also offer CSV download of updated data
                                csv_updated = updated_df.to_csv(index=False)
                                st.download_button(
                                    label="📥 Download Updated Data (CSV)",
                                    data=csv_updated,
                                    file_name="updated_data.csv",
                                    mime="text/csv",
                                    use_container_width=True
                                )

                            st.success("✅ Auto-fill process completed!")
                            st.info(f"""
                            **📊 Excel Report Contains 3 Sheets:**
                            1. **Summary** - Overview of auto-fill results
                            2. **Search Results** - Detailed search findings with reasons
                            3. **Updated Data** - Your original data with newly filled information

                            **APIs Used:** Google Custom Search, Hunter.io
                            """)

                else:
                    st.warning("No relevant columns found. Please ensure your CSV has name, email, or phone columns.")
                    st.info(f"Available columns: {', '.join(df.columns.tolist())}")

            except Exception as e:
                st.error(f"Error in auto-fill process: {str(e)}")
        else:
            st.info("Please upload a CSV file first to use the auto-fill feature.")

    elif st.session_state['feature'] == 'increase_reach':
        st.header("📈 Increase Reach")
        st.info("Find new leads by location and profession using web scraping from LinkedIn and other sources.")

        # Define dropdown options
        indian_cities = [
            "Mumbai", "Delhi", "Bangalore", "Hyderabad", "Ahmedabad", "Chennai", "Kolkata", 
            "Surat", "Pune", "Jaipur", "Lucknow", "Kanpur", "Nagpur", "Indore", "Thane",
        ]

        professions = [
            "Architects", "HVAC Consultants", "Industries", "Factories", "Data Centres", 
            "Corporate Offices", "Warehouses", "Builders and Developers"
        ]

        # Configuration
        st.subheader("⚙️ Search Configuration")
        col1, col2 = st.columns(2)

        with col1:
            selected_location = st.selectbox("Select Location:", indian_cities)
            selected_profession = st.selectbox("Select Profession:", professions)

        with col2:
            num_leads = st.slider("Number of leads to find:", min_value=5, max_value=50, value=20)
            search_depth = st.selectbox("Search Depth:", ["Basic", "Deep", "Comprehensive"])

        
        api_col1, api_col2 = st.columns(2)

        with api_col1:
            # Google Custom Search - Free tier
            st.subheader("🔑 API Configuration")
            st.info("**Required for real searches.** Get your credentials from Google Cloud Console.")
            
            google_api_key = st.text_input(
                "Google Custom Search API Key *", 
                value="AIzaSyBIlvP6tk8HDTy1F3qZ-Ir0xi8RiRHHfec",
                type="password", 
                help="Free tier: 100 searches/day. Get from console.cloud.google.com -> APIs & Services -> Credentials",
                placeholder="Enter your real Google API key"
            )
            google_search_engine_id = st.text_input(
                "Google Search Engine ID *", 
                value="902681f49f1d54896",
                help="Create custom search engine at programmablesearchengine.google.com/controlpanel/create",
                placeholder="Enter your Search Engine ID (format: xxxxxxxxx:yyyyyyy)"
            )

        st.divider()
        
        search_method = st.radio(
            "Search Method:",
            ["Real APIs (Production)", "Demo Mode (Simulation)"],
            index=0,
            help="Real APIs will use your Google credentials to fetch actual results. Demo Mode generates fake data for testing."
        )

        # Validation for Real APIs
        if search_method == "Real APIs (Production)":
            if not google_api_key or not google_search_engine_id:
                st.error("❌ **API credentials required!** Please enter both Google API Key and Search Engine ID above.")
            elif len(google_api_key) < 30:
                st.warning("⚠️ API Key seems too short. Make sure you entered the complete key.")
            else:
                st.success("✅ API credentials provided. Ready to search!")

        # Disable button if using Real APIs without credentials
        button_disabled = (search_method == "Real APIs (Production)" and (not google_api_key or not google_search_engine_id))
        
        if st.button("🔍 Find New Leads", disabled=button_disabled):
            # Additional validation
            if search_method == "Real APIs (Production)" and (not google_api_key or not google_search_engine_id):
                st.error("❌ Cannot proceed without API credentials. Please enter your Google API key and Search Engine ID.")
                st.stop()
            
            with st.spinner("Searching for leads..."):
                def simulate_lead_search(location, profession, num_leads):
                    """Simulate lead search results"""
                    leads = []
                    for i in range(num_leads):
                        # Generate realistic names
                        first_names = ["Rahul", "Priya", "Amit", "Sneha", "Vikram", "Anjali", "Rohit", "Kavita", "Arjun", "Meera"]
                        last_names = ["Sharma", "Patel", "Singh", "Kumar", "Gupta", "Jain", "Agarwal", "Verma", "Chopra", "Malhotra"]
                        
                        name = f"{np.random.choice(first_names)} {np.random.choice(last_names)}"
                        email = f"{name.lower().replace(' ', '.')}@example.com"
                        phone = f"+91-{np.random.randint(7000000000, 9999999999)}"
                        company = f"{profession} Solutions {np.random.randint(1, 100)}"
                        
                        lead = {
                            'Name': name,
                            'Profession': profession,
                            'Location': location,
                            'Email': email,
                            'Phone': phone,
                            'Company': company,
                            'Source': 'Simulated Web Search',
                            'Confidence': np.random.choice(['High', 'Medium', 'Low'], p=[0.4, 0.4, 0.2])
                        }
                        leads.append(lead)
                        time.sleep(0.01)  # Simulate search time
                    return leads

                def calculate_confidence_score(lead):
                    """Calculate confidence score for lead conversion (0-100)"""
                    score = 0
                    factors = []
                    
                    # Name quality (20 points)
                    if lead.get('Name') and lead['Name'] != 'N/A':
                        name = lead['Name'].strip()
                        # Filter out generic terms
                        generic_terms = ['contact', 'email', 'call', 'office', 'global', 'worldwide', 
                                       'design', 'architecture', 'institute', 'updation', 'us']
                        is_generic = any(term in name.lower() for term in generic_terms)
                        
                        if not is_generic and len(name.split()) >= 2:
                            score += 20
                            factors.append("✓ Valid person name")
                        elif not is_generic:
                            score += 10
                            factors.append("⚠ Single word name")
                    
                    # Email quality (30 points)
                    if lead.get('Email') and lead['Email'] != 'N/A':
                        email = lead['Email'].lower()
                        if '@' in email and '.' in email:
                            # Personal email domains get higher score
                            if any(domain in email for domain in ['gmail', 'yahoo', 'hotmail', 'outlook']):
                                score += 15
                                factors.append("⚠ Personal email domain")
                            else:
                                score += 30
                                factors.append("✓ Corporate email")
                    
                    # Phone quality (25 points)
                    if lead.get('Phone') and lead['Phone'] != 'N/A':
                        phone = re.sub(r'[^\d+]', '', lead['Phone'])
                        if len(phone) >= 10:
                            if phone.startswith('+91') or phone.startswith('91'):
                                score += 25
                                factors.append("✓ Valid Indian number")
                            else:
                                score += 20
                                factors.append("✓ Valid phone number")
                    
                    # Company information (15 points)
                    if lead.get('Company') and lead['Company'] != 'N/A':
                        score += 15
                        factors.append("✓ Company identified")
                    
                    # LinkedIn source bonus (10 points)
                    if lead.get('Source') and 'linkedin' in lead['Source'].lower():
                        score += 10
                        factors.append("✓ LinkedIn verified")
                    
                    return min(score, 100), factors

                def is_valid_person_name(name):
                    """Check if name is likely a real person, not a generic term"""
                    if not name or name == 'N/A':
                        return False
                    
                    name_lower = name.lower().strip()
                    
                    # Filter out generic/company terms
                    invalid_terms = [
                        'contact us', 'email', 'call us', 'phone', 'office', 'global', 
                        'worldwide', 'design', 'architecture', 'architect', 'institute',
                        'updation', 'login', 'signup', 'register', 'about', 'services',
                        'portfolio', 'projects', 'team', 'careers', 'join', 'follow',
                        'subscribe', 'newsletter', 'home', 'menu', 'search', 'headquarters'
                    ]
                    
                    if any(term in name_lower for term in invalid_terms):
                        return False
                    
                    # Must have at least 2 words (first and last name)
                    if len(name.split()) < 2:
                        return False
                    
                    # Should start with capital letter
                    if not name[0].isupper():
                        return False
                    
                    return True

                def search_google(location, profession, api_key=None, engine_id=None):
                    """Search Google for professionals - REAL API ONLY"""
                    if not api_key or not engine_id:
                        st.error("❌ Missing API key or Search Engine ID")
                        return []
                    
                    try:
                        leads = []
                        
                        # Multiple search strategies for better results
                        search_queries = [
                            f'{profession} {location} India site:linkedin.com/in',
                            f'{profession} {location} India email phone contact',
                            f'"{profession}" "{location}" director manager principal',
                            f'{profession} professionals {location} India contact details'
                        ]
                        
                        for query_idx, query in enumerate(search_queries):
                            st.info(f"🔍 Search strategy {query_idx + 1}/{len(search_queries)}: {query[:50]}...")
                            
                            url = f"https://www.googleapis.com/customsearch/v1?key={api_key}&cx={engine_id}&q={requests.utils.quote(query)}&num=10"
                            response = requests.get(url, timeout=15)
                            
                            # Check for API errors
                            if response.status_code != 200:
                                error_data = response.json() if response.text else {}
                                error_msg = error_data.get('error', {}).get('message', 'Unknown error')
                                st.warning(f"⚠️ Query {query_idx + 1} failed: {error_msg}")
                                continue
                            
                            data = response.json()
                            
                            if 'items' not in data:
                                continue
                            
                            for item in data.get('items', []):
                                title = item.get('title', '')
                                snippet = item.get('snippet', '')
                                link = item.get('link', '')
                                
                                # Multiple name extraction patterns
                                name = None
                                
                                # LinkedIn profile pattern
                                if 'linkedin.com' in link:
                                    linkedin_name = re.search(r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})\s*[-|]', title)
                                    if linkedin_name:
                                        name = linkedin_name.group(1).strip()
                                
                                # Directory listing pattern
                                if not name:
                                    dir_pattern = re.search(r'(?:Mr\.|Ms\.|Dr\.)?\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})', title)
                                    if dir_pattern:
                                        potential_name = dir_pattern.group(1).strip()
                                        if is_valid_person_name(potential_name):
                                            name = potential_name
                                
                                # Snippet pattern
                                if not name:
                                    snippet_pattern = re.search(r'(?:by|from|contact|reach)\s+([A-Z][a-z]+\s+[A-Z][a-z]+)', snippet)
                                    if snippet_pattern:
                                        potential_name = snippet_pattern.group(1).strip()
                                        if is_valid_person_name(potential_name):
                                            name = potential_name
                                
                                # Skip if no valid name found
                                if not name or not is_valid_person_name(name):
                                    continue
                                
                                # Extract email - improved patterns
                                email = None
                                email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
                                email_matches = re.findall(email_pattern, snippet + ' ' + title)
                                
                                for em in email_matches:
                                    # Skip generic emails
                                    if not any(generic in em.lower() for generic in ['info@', 'contact@', 'admin@', 'support@']):
                                        email = em
                                        break
                                
                                # Extract phone - improved for Indian numbers
                                phone = None
                                phone_patterns = [
                                    r'\+91[-\s]?\d{10}',
                                    r'91[-\s]?\d{10}',
                                    r'\d{5}[-\s]?\d{5}',
                                    r'\+91[-\s]?\d{5}[-\s]?\d{5}'
                                ]
                                
                                for pattern in phone_patterns:
                                    phone_match = re.search(pattern, snippet + ' ' + title)
                                    if phone_match:
                                        phone = phone_match.group(0)
                                        break
                                
                                # Extract company
                                company = None
                                company_patterns = [
                                    r'(?:at|@|works at|employed at|with)\s+([A-Z][a-zA-Z\s&]+?)(?:\s*[-|,]|\s*$)',
                                    r'([A-Z][a-zA-Z\s&]+?)\s*[-|]',
                                ]
                                
                                for pattern in company_patterns:
                                    company_match = re.search(pattern, snippet)
                                    if company_match:
                                        potential_company = company_match.group(1).strip()
                                        if len(potential_company) > 3 and potential_company not in name:
                                            company = potential_company
                                            break
                                
                                # Create lead entry
                                lead = {
                                    'Name': name,
                                    'Profession': profession,
                                    'Location': location,
                                    'Email': email if email else 'N/A',
                                    'Phone': phone if phone else 'N/A',
                                    'Company': company if company else 'N/A',
                                    'Source': 'LinkedIn' if 'linkedin.com' in link else 'Google Search',
                                    'Link': link,
                                    'Confidence Score': 0,
                                    'Confidence Factors': ''
                                }
                                
                                # Calculate confidence score
                                score, factors = calculate_confidence_score(lead)
                                lead['Confidence Score'] = score
                                lead['Confidence Factors'] = ' | '.join(factors)
                                
                                # Only add if confidence score is reasonable
                                if score >= 30:
                                    leads.append(lead)
                            
                            # Small delay between queries to avoid rate limiting
                            time.sleep(1)
                        
                        # Remove duplicates based on name
                        seen_names = set()
                        unique_leads = []
                        for lead in leads:
                            if lead['Name'] not in seen_names:
                                seen_names.add(lead['Name'])
                                unique_leads.append(lead)
                        
                        # Sort by confidence score
                        unique_leads.sort(key=lambda x: x['Confidence Score'], reverse=True)
                        
                        if unique_leads:
                            st.success(f"✅ Found {len(unique_leads)} qualified leads with confidence scores")
                        else:
                            st.warning("⚠️ No qualified leads found. Try adjusting search criteria.")
                        
                        return unique_leads
                        
                    except requests.exceptions.Timeout:
                        st.error("❌ API request timed out. Please try again.")
                        return []
                    except requests.exceptions.RequestException as e:
                        st.error(f"❌ Network error: {str(e)}")
                        return []
                    except Exception as e:
                        st.error(f"❌ Google Search error: {str(e)}")
                    
                    return []

                # Perform search
                if search_method == "Demo Mode (Simulation)":
                    new_leads = simulate_lead_search(selected_location, selected_profession, num_leads)
                else:
                    new_leads = []
                    
                    # Try Google - Real API call only
                    if google_api_key and google_search_engine_id:
                        st.info(f"🔍 Calling Google Custom Search API for '{selected_profession}' in '{selected_location}'...")
                        google_leads = search_google(selected_location, selected_profession, google_api_key, google_search_engine_id)
                        new_leads.extend(google_leads)
                    
                    # Show error if no real results - DO NOT fall back to simulation
                    if not new_leads:
                        st.error("❌ No leads found using Google Custom Search API. This could be due to:")
                        st.warning("""
                        - Invalid or expired API credentials
                        - No matching results for your search criteria
                        - API quota exceeded (100 queries/day for free tier)
                        - Search Engine ID not configured properly
                        
                        **Please verify your API credentials or try different search criteria.**
                        """)
                        st.stop()  # Stop execution, don't show dummy data

                if new_leads:
                    leads_df = pd.DataFrame(new_leads)
                    
                    # Summary with confidence metrics
                    st.subheader("📊 Search Summary")
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Total Leads Found", len(new_leads))
                    with col2:
                        avg_score = sum(l['Confidence Score'] for l in new_leads) / len(new_leads) if new_leads else 0
                        st.metric("Avg Confidence", f"{avg_score:.0f}%")
                    with col3:
                        high_quality = sum(1 for l in new_leads if l['Confidence Score'] >= 70)
                        st.metric("High Quality (70+)", high_quality)
                    with col4:
                        with_contact = sum(1 for l in new_leads if l['Email'] != 'N/A' or l['Phone'] != 'N/A')
                        st.metric("With Contact Info", with_contact)
                    
                    # Show top 5 highest confidence leads
                    st.subheader("🎯 Top Conversion Candidates")
                    st.info("These leads have the highest likelihood of conversion based on data quality")
                    
                    top_leads = sorted(new_leads, key=lambda x: x['Confidence Score'], reverse=True)[:5]
                    for i, lead in enumerate(top_leads, 1):
                        with st.expander(f"#{i} - {lead['Name']} (Score: {lead['Confidence Score']}%)", expanded=(i<=3)):
                            col_a, col_b = st.columns(2)
                            with col_a:
                                st.write(f"**📧 Email:** {lead['Email']}")
                                st.write(f"**📞 Phone:** {lead['Phone']}")
                                st.write(f"**🏢 Company:** {lead['Company']}")
                            with col_b:
                                st.write(f"**📍 Location:** {lead['Location']}")
                                st.write(f"**💼 Profession:** {lead['Profession']}")
                                st.write(f"**🔗 Source:** {lead['Source']}")
                            
                            # Show confidence factors
                            if lead.get('Confidence Factors'):
                                st.success(f"**Why this lead?** {lead['Confidence Factors']}")
                    
                    st.divider()
                    
                    # Full leads table
                    st.subheader("📋 All Leads (Sorted by Confidence)")
                    
                    # Color code by confidence score
                    def color_confidence(val):
                        if isinstance(val, (int, float)):
                            if val >= 70:
                                return 'background-color: #d4edda'
                            elif val >= 50:
                                return 'background-color: #fff3cd'
                            else:
                                return 'background-color: #f8d7da'
                        return ''
                    
                    styled_df = leads_df.style.applymap(color_confidence, subset=['Confidence Score'])
                    st.dataframe(styled_df, use_container_width=True)

                    # Download options
                    csv_leads = leads_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download Leads (CSV)",
                        data=csv_leads,
                        file_name=f"{selected_profession.lower().replace(' ', '_')}_{selected_location.lower()}_leads.csv",
                        mime="text/csv"
                    )

                    # Create Excel file for download
                    from io import BytesIO
                    from openpyxl.styles import PatternFill, Font

                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "New Leads"

                    # Add summary
                    ws['A1'] = f"Leads Search Results: {selected_profession} in {selected_location}"
                    ws['A1'].font = Font(bold=True, size=14)
                    ws.merge_cells('A1:E1')
                    
                    ws['A3'] = "Search Criteria:"
                    ws['B3'] = f"{selected_profession} in {selected_location}"
                    ws['A4'] = "Total Leads:"
                    ws['B4'] = len(new_leads)
                    ws['A5'] = "Search Method:"
                    ws['B5'] = search_method

                    # Add headers starting from row 7
                    headers = list(leads_df.columns)
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=7, column=col_num, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.alignment = Alignment(horizontal="center")

                    # Add data
                    for row_num, row_data in enumerate(leads_df.values, 8):
                        for col_num, value in enumerate(row_data, 1):
                            ws.cell(row=row_num, column=col_num, value=str(value))

                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = openpyxl.utils.get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width

                    # Save to BytesIO
                    excel_buffer = BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)

                    # Download button
                    st.download_button(
                        label="📥 Download Leads (Excel)",
                        data=excel_buffer,
                        file_name=f"{selected_profession.lower().replace(' ', '_')}_{selected_location.lower()}_leads.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                else:
                    st.info("No leads found for the selected criteria. Try different search parameters or API keys.")

    # Clear feature button
    if st.button("🔄 Back to Features"):
        del st.session_state['feature']
